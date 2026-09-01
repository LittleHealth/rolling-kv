"""Validate coverage across every experiment block and the artifact contracts."""

from __future__ import annotations

import json
import random
from collections import Counter
from pathlib import Path
from typing import Any

import numpy as np

from common import DATASETS, MODELS, RESULTS, read_jsonl, utc_now, write_json_atomic
from exp3_adaptive import policies


TERMINAL = {"ok", "unsupported", "oom", "capture_failed", "failed"}
HARD_FAILURE = {"capture_failed", "failed"}


def latest(path: Path, fields: tuple[str, ...]) -> dict[tuple[Any, ...], dict[str, Any]]:
    result = {}
    for row in read_jsonl(path):
        result[tuple(row.get(field) for field in fields)] = row
    return result


def exp3_lengths(model: str) -> tuple[int, ...]:
    spec = MODELS[model]
    return spec.lengths if spec.wave == "W1" else (spec.main_length,)


def validate() -> dict[str, Any]:
    failures: list[str] = []
    counts: Counter[str] = Counter()
    manifest_path = RESULTS / "manifest.json"
    if not manifest_path.exists():
        failures.append("manifest.json missing")
        manifest = {}
    else:
        manifest = json.loads(manifest_path.read_text())
    if manifest.get("other_processes_on_gpu"):
        failures.append("manifest other_processes_on_gpu is not empty")
    windows = manifest.get("windows", {})
    if any(len(windows.get(dataset, [])) != 5 for dataset in DATASETS):
        failures.append("manifest does not contain five shared windows per dataset")

    # EXP-0 gates.
    for model, spec in MODELS.items():
        rows = latest(
            RESULTS / "EXP0_correctness" / model / "records.jsonl",
            ("L", "gate", "cache_age"),
        )
        for length in spec.lengths:
            for gate in ("T1", "T2", "T3", "T4"):
                row = rows.get((length, gate, None))
                if not row or row.get("status") != "ok" or row.get("passed") is not True:
                    failures.append(f"EXP0 gate failed/missing: {model} L={length} {gate}")
            for age in range(1, 9):
                if rows.get((length, "T5", age), {}).get("status") != "ok":
                    failures.append(f"EXP0 T5 missing: {model} L={length} age={age}")

    # EXP-1 quality/timing complete grid.
    timing_inconsistent = timing_ok = 0
    npz_candidates: list[tuple[Path, dict[str, Any]]] = []
    for model, spec in MODELS.items():
        quality = latest(
            RESULTS / "EXP1_sweep" / model / "records.jsonl",
            ("dataset", "window", "L", "K", "pos_remap"),
        )
        timing = latest(
            RESULTS / "EXP1_sweep" / model / "timing.jsonl",
            ("L", "K", "pos_remap", "exec", "batch"),
        )
        for length in spec.lengths:
            for remap in spec.remap_values:
                for k in spec.k_values:
                    if timing.get((length, k, remap, "graph", 1), {}).get("status") not in TERMINAL:
                        failures.append(f"EXP1 timing missing: {model} L={length} K={k} remap={remap}")
                    row = timing.get((length, k, remap, "graph", 1))
                    if row and row.get("status") in HARD_FAILURE:
                        failures.append(
                            f"EXP1 timing failed: {model} L={length} K={k} remap={remap}"
                        )
                    if row and row.get("status") == "ok":
                        timing_ok += 1
                        timing_inconsistent += int(row.get("timing_flag") == "inconsistent")
                for dataset in DATASETS:
                    for window in range(5):
                        for k in spec.k_values:
                            key = (dataset, window, length, k, remap)
                            row = quality.get(key)
                            if not row or row.get("status") not in TERMINAL:
                                failures.append(f"EXP1 quality missing: {model} {key}")
                                continue
                            if row.get("status") in HARD_FAILURE:
                                failures.append(f"EXP1 quality failed: {model} {key}")
                            if row.get("status") == "ok" and k == 1:
                                for field, expected in (
                                    ("gap_pct", 0.0), ("mae_delta_pct", 0.0),
                                    ("speedup", 1.0),
                                ):
                                    if abs(float(row.get(field, float("inf"))) - expected) > 1e-6:
                                        failures.append(f"bad K1 {field}: {model} {key}")
                                pred_path = RESULTS / row["preds_file"]
                                if not pred_path.exists():
                                    failures.append(f"K1 NPZ missing: {pred_path}")
                                else:
                                    with np.load(pred_path, allow_pickle=False) as archive:
                                        if archive["yhat"].shape != (spec.updates, spec.horizon):
                                            failures.append(f"K1 NPZ shape bad: {pred_path}")
                                npz_candidates.append((pred_path, row))
    if timing_ok and timing_inconsistent / timing_ok > 0.05:
        failures.append(
            f"EXP1 timing inconsistent ratio {timing_inconsistent}/{timing_ok} exceeds 5%"
        )

    # EXP-2 main grid and consistency.
    stage_ok = stage_consistent = 0
    for model, spec in MODELS.items():
        rows = latest(
            RESULTS / "EXP2_stages" / model / "stages.jsonl",
            ("L", "path", "batch", "method"),
        )
        for length in spec.lengths:
            for path_name in ("full", "rolling"):
                row = rows.get((length, path_name, 1, "profiler_eager"))
                if not row or row.get("status") not in TERMINAL:
                    failures.append(f"EXP2 missing: {model} L={length} {path_name}")
                elif row.get("status") in HARD_FAILURE:
                    failures.append(f"EXP2 failed: {model} L={length} {path_name}")
                elif row.get("status") == "ok":
                    stage_ok += 1
                    stage_consistent += int(float(row.get("consistency_pct", 1e9)) <= 15.0)
                    if not row.get("stage_map"):
                        failures.append(f"EXP2 empty stage_map: {model} L={length} {path_name}")
        for path_name in ("full", "rolling"):
            batch8 = rows.get((spec.main_length, path_name, 8, "profiler_eager"), {})
            if batch8.get("status") not in TERMINAL:
                failures.append(f"EXP2 batch=8 missing: {model} {path_name}")
            elif batch8.get("status") in HARD_FAILURE:
                failures.append(f"EXP2 batch=8 failed: {model} {path_name}")
    if stage_ok and stage_consistent / stage_ok < 0.90:
        failures.append(f"EXP2 consistency pass ratio {stage_consistent}/{stage_ok} below 90%")

    # EXP-3 records, timing, traces and immutable per-policy prediction files.
    for model, spec in MODELS.items():
        records = latest(
            RESULTS / "EXP3_adaptive" / model / "records.jsonl",
            ("dataset", "window", "L", "policy_id"),
        )
        timing = latest(
            RESULTS / "EXP3_adaptive" / model / "timing.jsonl",
            ("dataset", "window", "L", "policy_id"),
        )
        for length in exp3_lengths(model):
            ids = list(policies(model, length))
            for dataset in DATASETS:
                for window in range(5):
                    for policy_id, _payload in ids:
                        key = (dataset, window, length, policy_id)
                        row = records.get(key)
                        timed = timing.get(key)
                        if not row or row.get("status") not in TERMINAL:
                            failures.append(f"EXP3 traced missing: {model} {key}")
                            continue
                        if row.get("status") in HARD_FAILURE:
                            failures.append(f"EXP3 traced failed: {model} {key}")
                        if not timed or timed.get("status") not in TERMINAL:
                            failures.append(f"EXP3 timed missing: {model} {key}")
                        elif timed.get("status") in HARD_FAILURE:
                            failures.append(f"EXP3 timed failed: {model} {key}")
                        if row.get("status") == "ok":
                            trace = latest(RESULTS / row["trace_file"], ("step",))
                            if set(index[0] for index in trace) != set(range(spec.updates)):
                                failures.append(f"EXP3 trace incomplete: {model} {key}")
                            if not (RESULTS / row["preds_file"]).exists():
                                failures.append(f"EXP3 predictions missing: {model} {key}")

    # EXP-5 required appendix grid.
    for model, spec in MODELS.items():
        rows = latest(
            RESULTS / "EXP5_appendix" / model / "eager_graph.jsonl",
            ("L", "path", "exec", "batch"),
        )
        for path_name in ("full", "rolling"):
            for execution in ("eager", "graph"):
                eager_graph = rows.get((spec.main_length, path_name, execution, 1), {})
                if eager_graph.get("status") not in TERMINAL:
                    failures.append(f"EXP5 eager/graph missing: {model} {path_name}/{execution}")
                elif eager_graph.get("status") in HARD_FAILURE:
                    failures.append(f"EXP5 eager/graph failed: {model} {path_name}/{execution}")
            for batch in (1, 4, 8, 16, 32):
                batch_row = rows.get((spec.main_length, path_name, "graph", batch), {})
                if batch_row.get("status") not in TERMINAL:
                    failures.append(f"EXP5 batch missing: {model} {path_name} b={batch}")
                elif batch_row.get("status") in HARD_FAILURE:
                    failures.append(f"EXP5 batch failed: {model} {path_name} b={batch}")
    timeflow = latest(
        RESULTS / "EXP5_appendix" / "sundial" / "timeflow.jsonl",
        ("sampling_steps", "K", "window"),
    )
    for steps in (1, 2, 5, 10, 20, 50):
        for k in (1, 4, 16):
            row = timeflow.get((steps, k, 0), {})
            if row.get("status") not in TERMINAL:
                failures.append(f"EXP5 Sundial TimeFlow missing: steps={steps} K={k}")
            elif row.get("status") in HARD_FAILURE:
                failures.append(f"EXP5 Sundial TimeFlow failed: steps={steps} K={k}")

    # Deterministic aggregate contract and required sheets.
    aggregate = RESULTS / "aggregate"
    workbook = aggregate / "rolling.xlsx"
    for required in (aggregate / "long.csv", workbook, aggregate / "aggregate_meta.json"):
        if not required.exists():
            failures.append(f"aggregate artifact missing: {required}")
    if workbook.exists():
        from openpyxl import load_workbook

        wb = load_workbook(workbook, read_only=True, data_only=True)
        required = {
            "overall-mae-x", "overall-mae-y", "overall-gap-x",
            "overall-gap-y", "overall-latency", "stages-main",
            "stale-speedup-x", "stale-speedup-y", "stale-worst-x",
            "stale-worst-y", "stale-mae-x", "stale-mae-y",
            "stale-gap-x", "stale-gap-y", "regret",
            "budget-compliance", "trigger-overhead",
        }
        required |= {f"pareto-{model}-x" for model in MODELS}
        required |= {f"pareto-{model}-y" for model in MODELS}
        required |= {f"oracle-tau-{model}" for model in MODELS}
        required |= {f"oracle-tau-time-{model}" for model in MODELS}
        required |= {f"ctx-mae-{model}" for model in MODELS}
        required |= {f"cacheage-gap-{model}" for model in MODELS}
        missing = sorted(required - set(wb.sheetnames))
        if missing:
            failures.append(f"aggregate sheets missing: {missing}")
        for name in wb.sheetnames:
            try:
                json.loads(wb[name]["A1"].value)
            except Exception:
                failures.append(f"sheet A1 is not JSON: {name}")

    # Recompute a deterministic sample of recorded metrics from NPZ files.
    rng = random.Random(20260819)
    for pred_path, row in rng.sample(npz_candidates, min(3, len(npz_candidates))):
        with np.load(pred_path, allow_pickle=False) as archive:
            mae = float(np.abs(archive["yhat"] - archive["y"]).mean())
        if abs(mae - float(row["mae_native"])) > 1e-6:
            failures.append(f"NPZ MAE mismatch: {pred_path}")

    counts.update(
        {
            "models": len(MODELS),
            "exp1_timing_ok": timing_ok,
            "exp1_timing_inconsistent": timing_inconsistent,
            "exp2_ok": stage_ok,
            "exp2_consistent": stage_consistent,
            "failure_count": len(failures),
        }
    )
    result = {
        "schema": "rolling-kv-validation",
        "ts": utc_now(),
        "passed": not failures,
        "counts": dict(counts),
        "failures": failures,
    }
    write_json_atomic(RESULTS / "validation_all.json", result)
    return result


def main() -> int:
    result = validate()
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["passed"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
