"""Deterministically aggregate JSONL records into long.csv and rolling.xlsx."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import tempfile
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd

from common import DATASETS, MODELS, RESULTS, json_safe, read_jsonl, write_json_atomic


def flatten(value: Any, prefix: str = "") -> dict[str, Any]:
    output: dict[str, Any] = {}
    if isinstance(value, dict):
        for key, item in value.items():
            child = f"{prefix}.{key}" if prefix else str(key)
            output.update(flatten(item, child))
    else:
        output[prefix] = value
    return output


def latest_rows(path: Path, fields: tuple[str, ...]) -> list[dict[str, Any]]:
    latest = {}
    for row in read_jsonl(path):
        latest[tuple(row.get(field) for field in fields)] = row
    return list(latest.values())


def load_all() -> dict[str, list[dict[str, Any]]]:
    data: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for model in MODELS:
        data["exp0"] += latest_rows(
            RESULTS / "EXP0_correctness" / model / "records.jsonl",
            ("model", "L", "gate", "cache_age"),
        )
        data["exp1"] += latest_rows(
            RESULTS / "EXP1_sweep" / model / "records.jsonl",
            ("model", "dataset", "window", "L", "K", "pos_remap"),
        )
        data["timing1"] += latest_rows(
            RESULTS / "EXP1_sweep" / model / "timing.jsonl",
            ("model", "L", "K", "pos_remap", "exec", "batch"),
        )
        data["exp2"] += latest_rows(
            RESULTS / "EXP2_stages" / model / "stages.jsonl",
            ("model", "L", "path", "batch", "method"),
        )
        data["exp3"] += latest_rows(
            RESULTS / "EXP3_adaptive" / model / "records.jsonl",
            ("model", "dataset", "window", "L", "policy_id"),
        )
        data["timing3"] += latest_rows(
            RESULTS / "EXP3_adaptive" / model / "timing.jsonl",
            ("model", "dataset", "window", "L", "policy_id"),
        )
    return data


def ok(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if row.get("status") == "ok"]


def median(rows: Iterable[dict[str, Any]], field: str) -> float | None:
    values = [float(row[field]) for row in rows if row.get(field) is not None]
    return float(np.median(values)) if values else None


def worst(rows: Iterable[dict[str, Any]], field: str) -> float | None:
    values = [float(row[field]) for row in rows if row.get(field) is not None]
    return max(values) if values else None


def excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(json_safe(value), ensure_ascii=False, sort_keys=True)
    return value


class SheetBook:
    def __init__(self) -> None:
        from openpyxl import Workbook

        self.book = Workbook()
        self.book.remove(self.book.active)
        self.book.properties.creator = "rolling-kv"
        self.book.properties.created = datetime(2026, 8, 19, tzinfo=timezone.utc)
        self.book.properties.modified = datetime(2026, 8, 19, tzinfo=timezone.utc)

    def add(
        self,
        name: str,
        header: list[Any],
        rows: Iterable[list[Any]],
        config: dict[str, Any],
    ) -> None:
        sheet = self.book.create_sheet(name[:31])
        sheet.cell(1, 1, json.dumps(config, ensure_ascii=False, sort_keys=True))
        for col, value in enumerate(header, 1):
            sheet.cell(2, col, excel_value(value))
        for row_index, values in enumerate(rows, 3):
            for col, value in enumerate(values, 1):
                sheet.cell(row_index, col, excel_value(value))
        sheet.freeze_panes = "B3"


def deterministic_xlsx(book: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(dir=path.parent) as tmpdir:
        raw = Path(tmpdir) / "raw.xlsx"
        normalized = Path(tmpdir) / "normalized.xlsx"
        book.save(raw)
        with zipfile.ZipFile(raw, "r") as source, zipfile.ZipFile(
            normalized, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target:
            for name in sorted(source.namelist()):
                info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = 0o600 << 16
                target.writestr(info, source.read(name))
        os.replace(normalized, path)


def add_overall(book: SheetBook, data: dict[str, list[dict[str, Any]]]) -> None:
    for metric, stem in (("mae_delta_pct", "mae"), ("gap_pct", "gap")):
        x_rows, y_rows = [], []
        for model, spec in MODELS.items():
            fixed = [
                row for row in ok(data["exp1"])
                if row["model"] == model and row["L"] == spec.main_length
                and row.get("pos_remap") == spec.pos_remap
            ]
            adaptive = [
                row for row in ok(data["exp3"])
                if row["model"] == model and row["L"] == spec.main_length
                and row.get("adaptive_params", {}).get("budget_pct") == 5.0
                and row.get("adaptive_params", {}).get("calib") == "per_model"
            ]
            for label, rows in (
                ("Full", [row for row in fixed if row.get("K") == 1]),
                ("Naive", [row for row in fixed if row.get("K") == 0]),
                ("Ours(b=5%)", adaptive),
            ):
                x_rows.append([f"{model}:{label}", median(rows, "speedup"), median(rows, "speedup")])
                y_rows.append([f"{model}:{label}", median(rows, metric), worst(rows, metric)])
        config = {"xlabel": "speedup", "ylabel": metric, "title": f"overall {stem}", "ylog": False, "hline": 0, "note": "median and worst over windows/datasets"}
        book.add(f"overall-{stem}-x", ["series", "median", "worst"], x_rows, config)
        book.add(f"overall-{stem}-y", ["series", "median", "worst"], y_rows, config)


def add_latency(book: SheetBook, data: dict[str, list[dict[str, Any]]]) -> None:
    rows = []
    for model, spec in MODELS.items():
        candidates = [
            row for row in ok(data["timing1"])
            if row["model"] == model and row["L"] == spec.main_length
            and row.get("K") == 1 and row.get("pos_remap") == spec.pos_remap
        ]
        row = candidates[-1] if candidates else {}
        full = row.get("t_full_ms", {}).get("median") if row else None
        roll = row.get("t_roll_ms", {}).get("median") if row else None
        rows.append([model, None if full is None else full / spec.s * 1000, None if roll is None else roll / spec.s * 1000])
    book.add("overall-latency", ["model", "full", "rolling"], rows, {"xlabel": "model", "ylabel": "us/point", "title": "per-point latency", "ylog": False, "hline": None, "note": "main context"})


def add_stages(book: SheetBook, data: dict[str, list[dict[str, Any]]]) -> None:
    stage_names = [f"S{i}_{name}" for i, name in enumerate(("norm", "embed", "attn", "ffn", "head", "cache", "other"), 1)]
    main_rows = []
    for stage in stage_names:
        values = [stage]
        for model, spec in MODELS.items():
            for path_name in ("full", "rolling"):
                rows = [row for row in ok(data["exp2"]) if row["model"] == model and row["L"] == spec.main_length and row["path"] == path_name and row["batch"] == 1]
                values.append(median([{"v": row.get("stage_busy_ms", {}).get(stage)} for row in rows], "v"))
        main_rows.append(values)
    header = ["stage"] + [f"{model}:{path}" for model in MODELS for path in ("full", "rolling")]
    book.add("stages-main", header, main_rows, {"xlabel": "model/path", "ylabel": "ms", "title": "stage breakdown", "ylog": False, "hline": None, "note": "profiler eager device busy time"})
    for model, spec in MODELS.items():
        rows = []
        for stage in stage_names:
            for path_name in ("full", "rolling"):
                values = [f"{stage}:{path_name}"]
                for length in spec.lengths:
                    candidates = [row for row in ok(data["exp2"]) if row["model"] == model and row["L"] == length and row["path"] == path_name and row["batch"] == 1]
                    values.append(median([{"v": row.get("stage_busy_ms", {}).get(stage)} for row in candidates], "v"))
                rows.append(values)
        book.add(f"stages-vs-L-{model}", ["series", *spec.lengths], rows, {"xlabel": "L", "ylabel": "ms", "title": f"{model} stages vs L", "ylog": False, "hline": None, "note": "absolute profiler time"})


def fixed_by_model(data: dict[str, list[dict[str, Any]]], model: str, length: int) -> list[dict[str, Any]]:
    spec = MODELS[model]
    return [row for row in ok(data["exp1"]) if row["model"] == model and row["L"] == length and row.get("pos_remap") == spec.pos_remap]


def add_stale_and_pareto(book: SheetBook, data: dict[str, list[dict[str, Any]]]) -> None:
    max_points = max(len(spec.k_values) for spec in MODELS.values())
    for field, stem in (("speedup", "speedup"), ("mae_delta_pct", "worst"), ("mae_native", "mae"), ("gap_pct", "gap")):
        x_rows, y_rows = [], []
        for model, spec in MODELS.items():
            grouped = defaultdict(list)
            for row in fixed_by_model(data, model, spec.main_length):
                grouped[int(row["K"])].append(row)
            ordered = [k for k in spec.k_values if k in grouped]
            x = [None if k == 0 else k * spec.s for k in ordered]
            reducer = worst if stem == "worst" else median
            y = [reducer(grouped[k], field) for k in ordered]
            x_rows.append([model, *x, *([None] * (max_points - len(x)))])
            y_rows.append([model, *y, *([None] * (max_points - len(y)))])
        header = ["series", *range(max_points)]
        cfg = {"xlabel": "tau (points)", "ylabel": field, "title": f"staleness {stem}", "ylog": False, "hline": 0, "note": "main L; median unless named worst"}
        book.add(f"stale-{stem}-x", header, x_rows, cfg)
        book.add(f"stale-{stem}-y", header, y_rows, cfg)

    for model, spec in MODELS.items():
        x_rows, y_rows = [], []
        for length in spec.lengths:
            grouped = defaultdict(list)
            for row in fixed_by_model(data, model, length):
                grouped[int(row["K"])].append(row)
            x_rows.append([f"L={length}", *[median(grouped.get(k, []), "speedup") for k in spec.k_values]])
            y_rows.append([f"L={length}", *[median(grouped.get(k, []), "gap_pct") for k in spec.k_values]])
        cfg = {"xlabel": "speedup", "ylabel": "gap_pct", "title": f"{model} Pareto", "ylog": False, "hline": 0, "note": "columns are K"}
        book.add(f"pareto-{model}-x", ["L", *spec.k_values], x_rows, cfg)
        book.add(f"pareto-{model}-y", ["L", *spec.k_values], y_rows, cfg)


def oracle_tau(rows: list[dict[str, Any]], spec: Any, budget: float) -> int | None:
    grouped = defaultdict(list)
    for row in rows:
        grouped[int(row["K"])].append(row)
    safe = []
    for k, values in grouped.items():
        if k > 0 and worst(values, "mae_delta_pct") is not None and worst(values, "mae_delta_pct") <= budget:
            safe.append((median(values, "speedup") or 0.0, k))
    return max(safe, default=(None, None))[1] * spec.s if safe else None


def add_oracles(book: SheetBook, data: dict[str, list[dict[str, Any]]]) -> None:
    for model, spec in MODELS.items():
        fixed = fixed_by_model(data, model, spec.main_length)
        rows = []
        for budget in BUDGETS:
            rows.append([budget, *[oracle_tau([row for row in fixed if row["dataset"] == dataset], spec, budget) for dataset in DATASETS]])
        book.add(f"oracle-tau-{model}", ["budget", *DATASETS], rows, {"xlabel": "dataset", "ylabel": "tau", "title": f"{model} oracle tau", "ylog": False, "hline": None, "note": "worst-window budget"})
        time_rows = []
        for dataset in DATASETS:
            time_rows.append([dataset, *[oracle_tau([row for row in fixed if row["dataset"] == dataset and row["window"] == window], spec, 5.0) for window in range(5)]])
        book.add(f"oracle-tau-time-{model}", ["dataset", *range(5)], time_rows, {"xlabel": "window", "ylabel": "tau", "title": f"{model} oracle tau over time", "ylog": False, "hline": None, "note": "b=5%"})


BUDGETS = (1.0, 2.0, 5.0, 10.0)


def add_adaptive(book: SheetBook, data: dict[str, list[dict[str, Any]]]) -> None:
    for metric in ("speedup", "mae_delta_pct"):
        rows = []
        for model, spec in MODELS.items():
            candidates = [row for row in ok(data["exp3"]) if row["model"] == model and row["L"] == spec.main_length and row.get("adaptive_params", {}).get("budget_pct") == 5.0 and row.get("adaptive_params", {}).get("calib") == "per_model"]
            rows.append([model, *[median([row for row in candidates if row["adaptive_params"]["theta"] == theta], metric) for theta in (0.005, 0.01, 0.02, 0.05, 0.1, 0.2, 0.5)]])
        book.add(f"threshold-{metric}", ["model", 0.005, 0.01, 0.02, 0.05, 0.1, 0.2, 0.5], rows, {"xlabel": "theta", "ylabel": metric, "title": f"threshold {metric}", "ylog": False, "hline": 0, "note": "b=5%, per_model"})

    regret_rows, overhead_rows = [], []
    compliance_rows = []
    for model, spec in MODELS.items():
        fixed = fixed_by_model(data, model, spec.main_length)
        adaptive = [row for row in ok(data["exp3"]) if row["model"] == model and row["L"] == spec.main_length and row.get("adaptive_params", {}).get("budget_pct") == 5.0 and row.get("adaptive_params", {}).get("calib") == "per_model"]
        global_tau = oracle_tau(fixed, spec, 5.0)
        global_rows = [row for row in fixed if row.get("tau_pts") == global_tau]
        oracle_speeds = []
        for dataset in DATASETS:
            subset = [row for row in fixed if row["dataset"] == dataset]
            tau = oracle_tau(subset, spec, 5.0)
            oracle_speeds.append(median([row for row in subset if row.get("tau_pts") == tau], "speedup"))
        regret_rows.append([model, median(global_rows, "speedup"), median(adaptive, "speedup"), float(np.nanmedian([x for x in oracle_speeds if x is not None])) if any(x is not None for x in oracle_speeds) else None])
        timing = [row for row in ok(data["timing3"]) if row["model"] == model and row["L"] == spec.main_length]
        overhead = median(timing, "trigger_overhead_ms")
        measured = median([{"v": row.get("t_update_measured_ms", {}).get("mean")} for row in timing], "v")
        overhead_rows.append([model, overhead, None if overhead is None or measured is None else overhead / max(measured, 1e-12) * 100])
        compliance_rows.append([model, *[(worst([row for row in adaptive if row["dataset"] == dataset], "mae_delta_pct") or 0) - 5.0 for dataset in DATASETS]])
    book.add("regret", ["model", "global-fixed", "adaptive", "oracle"], regret_rows, {"xlabel": "model", "ylabel": "speedup", "title": "regret", "ylog": False, "hline": None, "note": "b=5%"})
    book.add("budget-compliance", ["model", *DATASETS], compliance_rows, {"xlabel": "dataset", "ylabel": "actual worst delta - budget", "title": "budget compliance", "ylog": False, "hline": 0, "note": "b=5%"})
    book.add("trigger-overhead", ["model", "ms", "percent"], overhead_rows, {"xlabel": "model", "ylabel": "overhead", "title": "trigger overhead", "ylog": False, "hline": 0, "note": "signal vs replay"})


def add_context_and_age(book: SheetBook, data: dict[str, list[dict[str, Any]]]) -> None:
    for model, spec in MODELS.items():
        fixed = [row for row in ok(data["exp1"]) if row["model"] == model and row.get("K") == 1 and row.get("pos_remap") == spec.pos_remap]
        book.add(f"ctx-mae-{model}", ["series", *spec.lengths], [["K=1", *[median([row for row in fixed if row["L"] == length], "mae_native") for length in spec.lengths]]], {"xlabel": "L", "ylabel": "MAE", "title": f"{model} context accuracy", "ylog": False, "hline": None, "note": "median"})
        t5 = [row for row in ok(data["exp0"]) if row["model"] == model and row.get("gate") == "T5"]
        rows = [[f"L={length}", *[median([row for row in t5 if row["L"] == length and row["cache_age"] == age], "gap_rel") for age in range(1, 9)]] for length in spec.lengths]
        book.add(f"cacheage-gap-{model}", ["L", *range(1, 9)], rows, {"xlabel": "cache age", "ylabel": "gap_rel", "title": f"{model} cache age gap", "ylog": False, "hline": 0, "note": "EXP0 T5"})


def build(output: Path) -> dict[str, Any]:
    data = load_all()
    aggregate = RESULTS / "aggregate"
    aggregate.mkdir(parents=True, exist_ok=True)
    long_rows = []
    for family, rows in data.items():
        for row in rows:
            flat = flatten(row)
            flat["record_family"] = family
            long_rows.append(flat)
    frame = pd.DataFrame(long_rows)
    frame = frame.reindex(sorted(frame.columns), axis=1)
    frame = frame.sort_values([column for column in ("record_family", "model", "dataset", "window", "L", "K", "policy_id") if column in frame], kind="mergesort", na_position="last")
    frame.to_csv(aggregate / "long.csv", index=False, lineterminator="\n")

    book = SheetBook()
    add_overall(book, data)
    add_latency(book, data)
    add_stages(book, data)
    add_stale_and_pareto(book, data)
    add_oracles(book, data)
    add_adaptive(book, data)
    add_context_and_age(book, data)
    deterministic_xlsx(book.book, output)

    status_counts = Counter()
    reasons = Counter()
    for rows in data.values():
        for row in rows:
            status_counts[row.get("status", "missing")] += 1
            if row.get("status") != "ok":
                reasons[row.get("reason") or "unspecified"] += 1
    meta = {
        "schema": "rolling-kv-aggregate",
        "status_counts": dict(status_counts),
        "non_ok_reasons": dict(reasons),
        "row_count": len(long_rows),
        "sheet_names": book.book.sheetnames,
        "xlsx_sha256": hashlib.sha256(output.read_bytes()).hexdigest(),
    }
    write_json_atomic(aggregate / "aggregate_meta.json", meta)
    return meta


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--verify-determinism", action="store_true")
    args = parser.parse_args()
    output = RESULTS / "aggregate" / "rolling.xlsx"
    meta = build(output)
    if args.verify_determinism:
        first = output.read_bytes()
        build(output)
        if output.read_bytes() != first:
            raise RuntimeError("rolling.xlsx is not byte-deterministic")
    print(json.dumps(meta, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
